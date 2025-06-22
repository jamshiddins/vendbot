"""
Сервис для генерации отчетов
"""
import io
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any

import pandas as pd
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from sqlalchemy.orm import selectinload

from backend.models import (
    User, UserRole, Machine, Hopper, 
    IngredientType, Inventory, Operation
)

logger = logging.getLogger(__name__)


class ReportService:
    """
    Сервис для генерации различных отчетов
    """
    
    def __init__(self, session: AsyncSession):
        self.session = session
    
    async def generate_users_report(
        self,
        active_only: bool = False,
        role_filter: Optional[str] = None
    ) -> io.BytesIO:
        """
        Генерирует Excel отчет по пользователям
        """
        # Формируем запрос
        stmt = select(User).options(selectinload(User.roles))
        
        if active_only:
            stmt = stmt.where(User.is_active == True)
        
        # Выполняем запрос
        result = await self.session.execute(stmt)
        users = result.scalars().all()
        
        # Фильтруем по роли если нужно
        if role_filter:
            users = [u for u in users if u.has_role(role_filter)]
        
        # Подготавливаем данные для DataFrame
        data = []
        for user in users:
            roles = ", ".join(sorted(user.role_names))
            
            data.append({
                "ID": user.id,
                "Telegram ID": user.telegram_id,
                "Username": f"@{user.username}" if user.username else "",
                "Полное имя": user.full_name,
                "Телефон": user.phone or "",
                "Роли": roles,
                "Статус": "Активен" if user.is_active else "Заблокирован",
                "Владелец": "Да" if user.is_owner else "Нет",
                "Дата регистрации": user.created_at,
                "Последнее обновление": user.updated_at
            })
        
        # Создаем DataFrame
        df = pd.DataFrame(data)
        
        # Создаем Excel файл в памяти
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Пользователи', index=False)
            
            # Получаем лист для форматирования
            worksheet = writer.sheets['Пользователи']
            
            # Автоширина колонок
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
        
        output.seek(0)
        return output
    
    async def generate_operations_report(
        self,
        date_from: datetime,
        date_to: datetime,
        user_id: Optional[int] = None,
        operation_type: Optional[str] = None
    ) -> io.BytesIO:
        """
        Генерирует Excel отчет по операциям
        """
        # Формируем запрос
        stmt = select(Operation).options(
            selectinload(Operation.user)
        ).where(
            and_(
                Operation.created_at >= date_from,
                Operation.created_at <= date_to
            )
        )
        
        if user_id:
            stmt = stmt.where(Operation.user_id == user_id)
        
        if operation_type:
            stmt = stmt.where(Operation.operation_type == operation_type)
        
        stmt = stmt.order_by(Operation.created_at.desc())
        
        # Выполняем запрос
        result = await self.session.execute(stmt)
        operations = result.scalars().all()
        
        # Подготавливаем данные
        data = []
        for op in operations:
            data.append({
                "ID": op.id,
                "Дата/время": op.created_at,
                "Пользователь": op.user.full_name if op.user else "Неизвестно",
                "Тип операции": op.display_type,
                "Объект": f"{op.entity_type or ''} {op.entity_id or ''}".strip(),
                "Описание": op.description,
                "Статус": "Успешно" if op.success else "Ошибка",
                "Ошибка": op.error_message or ""
            })
        
        # Создаем DataFrame
        df = pd.DataFrame(data)
        
        # Создаем Excel
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Операции', index=False)
            
            # Форматирование
            worksheet = writer.sheets['Операции']
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
        
        output.seek(0)
        return output
    
    async def generate_stock_report(self) -> io.BytesIO:
        """
        Генерирует Excel отчет по остаткам на складе
        """
        # Получаем данные
        stmt = select(IngredientType, Inventory).join(
            Inventory,
            IngredientType.id == Inventory.ingredient_type_id,
            isouter=True
        ).order_by(IngredientType.category, IngredientType.name)
        
        result = await self.session.execute(stmt)
        items = result.all()
        
        # Подготавливаем данные
        data = []
        for ingredient_type, inventory in items:
            quantity = inventory.quantity if inventory else 0
            reserved = inventory.reserved_quantity if inventory else 0
            available = quantity - reserved
            
            # Определяем статус
            if quantity <= ingredient_type.min_stock_level:
                status = "Критический"
            elif quantity <= ingredient_type.reorder_level:
                status = "Низкий"
            elif quantity >= ingredient_type.max_stock_level:
                status = "Избыток"
            else:
                status = "Нормальный"
            
            data.append({
                "Категория": ingredient_type.category,
                "Наименование": ingredient_type.name,
                "Ед. изм.": ingredient_type.unit,
                "Всего на складе": quantity,
                "Зарезервировано": reserved,
                "Доступно": available,
                "Мин. уровень": ingredient_type.min_stock_level,
                "Уровень заказа": ingredient_type.reorder_level,
                "Макс. уровень": ingredient_type.max_stock_level,
                "Статус": status,
                "Последнее пополнение": inventory.last_restock_date if inventory else None
            })
        
        # Создаем DataFrame
        df = pd.DataFrame(data)
        
        # Создаем Excel с форматированием
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Остатки на складе', index=False)
            
            # Форматирование
            worksheet = writer.sheets['Остатки на складе']
            
            # Автоширина колонок
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 40)
            
            # Условное форматирование для статусов
            from openpyxl.styles import PatternFill
            
            for row in range(2, len(df) + 2):  # +2 для заголовка и индексации с 1
                status_cell = worksheet[f'J{row}']  # Колонка "Статус"
                
                if status_cell.value == "Критический":
                    status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif status_cell.value == "Низкий":
                    status_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif status_cell.value == "Избыток":
                    status_cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        output.seek(0)
        return output
    
    async def get_stock_summary(self) -> Dict[str, Any]:
        """
        Получает сводку по складским остаткам
        """
        # Общее количество типов
        total_types = await self.session.scalar(
            select(func.count(IngredientType.id))
        )
        
        # Статистика по уровням
        stmt = select(
            IngredientType,
            Inventory
        ).join(
            Inventory,
            IngredientType.id == Inventory.ingredient_type_id,
            isouter=True
        )
        
        result = await self.session.execute(stmt)
        items = result.all()
        
        critical = 0
        low = 0
        normal = 0
        excess = 0
        
        categories_stats = {}
        
        for ingredient_type, inventory in items:
            quantity = inventory.quantity if inventory else 0
            
            # Статус
            if quantity <= ingredient_type.min_stock_level:
                critical += 1
            elif quantity <= ingredient_type.reorder_level:
                low += 1
            elif quantity >= ingredient_type.max_stock_level:
                excess += 1
            else:
                normal += 1
            
            # По категориям
            category = ingredient_type.category
            if category not in categories_stats:
                categories_stats[category] = {
                    "count": 0,
                    "total_quantity": 0,
                    "unit": ingredient_type.unit
                }
            
            categories_stats[category]["count"] += 1
            categories_stats[category]["total_quantity"] += quantity
        
        return {
            "summary": {
                "total_types": total_types,
                "critical": critical,
                "low": low,
                "normal": normal,
                "excess": excess
            },
            "by_category": categories_stats,
            "generated_at": datetime.now().isoformat()
        }
    
    async def generate_machine_report(self) -> io.BytesIO:
        """
        Генерирует отчет по автоматам
        """
        # Получаем все автоматы с операторами
        stmt = select(Machine).options(
            selectinload(Machine.assigned_operator),
            selectinload(Machine.hoppers)
        )
        
        result = await self.session.execute(stmt)
        machines = result.scalars().all()
        
        # Подготавливаем данные
        data = []
        for machine in machines:
            # Подсчитываем бункеры
            installed_hoppers = len([h for h in machine.hoppers if h.status == "installed"])
            
            data.append({
                "Код": machine.code,
                "Название": machine.name,
                "Адрес": machine.location_address,
                "Детали": machine.location_details or "",
                "Статус": machine.status,
                "Оператор": machine.assigned_operator.full_name if machine.assigned_operator else "Не назначен",
                "Бункеров установлено": installed_hoppers,
                "Дата установки": machine.installation_date,
                "Последнее обслуживание": machine.last_service_date,
                "Широта": machine.latitude or "",
                "Долгота": machine.longitude or ""
            })
        
        # Создаем DataFrame
        df = pd.DataFrame(data)
        
        # Создаем Excel
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Автоматы', index=False)
            
            # Форматирование
            worksheet = writer.sheets['Автоматы']
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
        
        output.seek(0)
        return output


# Экспорт
__all__ = ['ReportService']